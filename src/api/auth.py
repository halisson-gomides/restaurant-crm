"""Authentication API routes for admin login system."""

from datetime import datetime, date, time, timedelta
from fastapi import APIRouter, Depends, HTTPException, status, Request, Response, Query
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, literal
from typing import Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..database import get_database
from ..models.client_registration import User, CNPJRegistration, CPFRegistration, Organization
from ..schemas.auth import UserLogin, Token
from ..utils.security import create_access_token, verify_token
from ..api.deps import get_current_user

router = APIRouter(prefix="/auth", tags=["authentication"])
security = HTTPBearer()


def parse_date_filter(date_str: str) -> datetime:
    """Parse date string to datetime for filtering."""
    try:
        # Parse as date and create datetime at start/end of day
        parsed_date = date.fromisoformat(date_str)
        return datetime.combine(parsed_date, time.min)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid date format: {date_str}. Use YYYY-MM-DD format."
        )


@router.get("/logout")
async def logout():
    """Logout endpoint - clears the auth cookie and redirects."""
    response = RedirectResponse(url="/auth/login", status_code=302)
    response.delete_cookie(key="admin_token", path="/", httponly=True)
    return response


@router.post("/login")
async def login(
    user_data: UserLogin,
    response: Response,
    db: AsyncSession = Depends(get_database)
):
    """Authenticate admin user and return access token."""
    # Find user by username (CPF/CNPJ)
    result = await db.execute(
        select(User).where(User.username == user_data.username)
    )
    user = result.scalar_one_or_none()

    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Credenciais inválidas",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Verify password (using simple hash for now)
    import hashlib
    salt = "restaurant_crm_salt"
    hashed_password = hashlib.sha256(f"{user_data.password}{salt}".encode()).hexdigest()

    if user.hashed_password != hashed_password:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Credenciais inválidas",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Check if user is active and admin
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Conta desativada"
        )

    if user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado. Requer privilégio de administrador"
        )

    # Create access token
    access_token = create_access_token(
        data={"sub": user.username, "role": user.role},
        expires_delta=timedelta(minutes=60)  # 1 hour for admin sessions
    )

    # Set cookie for browser requests
    response.set_cookie(
        key="admin_token",
        value=access_token,
        httponly=True,  # Prevent JavaScript access for security
        secure=False,   # Set to True in production with HTTPS
        samesite="lax",
        path="/",       # Make cookie available for all paths
        max_age=3600    # 1 hour
    )

    return Token(access_token=access_token, token_type="bearer")


# @router.get("/me")
# async def get_current_user_info(
#     current_user: User = Depends(get_current_user)
# ):
#     """Get current authenticated user information."""
#     return {
#         "id": current_user.id,
#         "username": current_user.username,
#         "email": current_user.email,
#         "first_name": current_user.first_name,
#         "last_name": current_user.last_name,
#         "role": current_user.role,
#         "is_active": current_user.is_active
#     }


@router.get("/registrations/export")
async def export_registrations(
    request: Request,
    format: str = "excel",
    registration_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_database)
):
    """Export registrations data in specified format (admin only)."""
    # Try to get token from Authorization header first
    token = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
    else:
        # Fall back to cookie
        token = request.cookies.get("admin_token")

    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required"
        )

    # Verify token
    payload = verify_token(token)
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication token"
        )

    # Check if user is admin
    if payload.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    # Build queries for CNPJ and CPF registrations (no pagination for export)
    cnpj_query = select(
        # CNPJRegistration.id,
        CNPJRegistration.razao_social.label("name"),
        CNPJRegistration.cnpj.label("document"),
        CNPJRegistration.email,
        CNPJRegistration.celular.label("phone"),
        CNPJRegistration.created_at,
        literal("CNPJ").label("type"),
        CNPJRegistration.qual_seu_negocio.label("business_type"),
        CNPJRegistration.seu_nome.label("contact_name"),
        CNPJRegistration.sua_funcao.label("contact_role"),
        CNPJRegistration.cep,
        CNPJRegistration.endereco,
        CNPJRegistration.bairro,
        CNPJRegistration.cidade,
        CNPJRegistration.estado
    )

    cpf_query = select(
        # CPFRegistration.id,
        CPFRegistration.nome_completo.label("name"),
        CPFRegistration.cpf.label("document"),
        CPFRegistration.email,
        CPFRegistration.celular.label("phone"),
        CPFRegistration.created_at,
        literal("CPF").label("type"),
        CPFRegistration.perfil_compra.label("purchase_profile"),
        CPFRegistration.qual_negocio_cpf.label("business_name"),
        CPFRegistration.genero.label("gender"),
        CPFRegistration.data_nascimento.label("birth_date"),
        CPFRegistration.cep,
        CPFRegistration.endereco,
        CPFRegistration.bairro,
        CPFRegistration.cidade,
        CPFRegistration.estado
    )

    # Apply filters
    if registration_type:
        if registration_type == "CNPJ":
            cpf_query = None
        elif registration_type == "CPF":
            cnpj_query = None

    if date_from:
        date_from_dt = parse_date_filter(date_from)
        if cnpj_query is not None:
            cnpj_query = cnpj_query.where(CNPJRegistration.created_at >= date_from_dt)
        if cpf_query is not None:
            cpf_query = cpf_query.where(CPFRegistration.created_at >= date_from_dt)

    if date_to:
        date_to_dt = parse_date_filter(date_to)
        # Set to end of day for date_to
        date_to_dt = datetime.combine(date_to_dt.date(), time.max)
        if cnpj_query is not None:
            cnpj_query = cnpj_query.where(CNPJRegistration.created_at <= date_to_dt)
        if cpf_query is not None:
            cpf_query = cpf_query.where(CPFRegistration.created_at <= date_to_dt)

    if search:
        search_term = f"%{search}%"
        if cnpj_query is not None:
            cnpj_query = cnpj_query.where(
                (CNPJRegistration.razao_social.ilike(search_term)) |
                (CNPJRegistration.email.ilike(search_term))
            )
        if cpf_query is not None:
            cpf_query = cpf_query.where(
                (CPFRegistration.nome_completo.ilike(search_term)) |
                (CPFRegistration.email.ilike(search_term))
            )

    # Execute queries and combine results
    all_registrations = []

    if cnpj_query is not None:
        cnpj_result = await db.execute(cnpj_query)
        for row in cnpj_result:
            all_registrations.append({
                # "id": row.id,
                "type": row.type,
                "name": row.name,
                "document": row.document,
                "email": row.email,
                "phone": row.phone,
                "created_at": row.created_at.strftime("%d/%m/%Y %H:%M") if row.created_at else None,
                "business_type": row.business_type,
                "contact_name": row.contact_name,
                "contact_role": row.contact_role,
                "cep": row.cep,
                "endereco": row.endereco,
                "bairro": row.bairro,
                "cidade": row.cidade,
                "estado": row.estado,
                "purchase_profile": None,
                "business_name": None,
                "gender": None,
                "birth_date": None
            })

    if cpf_query is not None:
        cpf_result = await db.execute(cpf_query)
        for row in cpf_result:
            all_registrations.append({
                # "id": row.id,
                "type": row.type,
                "name": row.name,
                "document": row.document,
                "email": row.email,
                "phone": row.phone,
                "created_at": row.created_at.strftime("%d/%m/%Y %H:%M") if row.created_at else None,
                "business_type": None,
                "contact_name": None,
                "contact_role": None,
                "cep": row.cep,
                "endereco": row.endereco,
                "bairro": row.bairro,
                "cidade": row.cidade,
                "estado": row.estado,
                "purchase_profile": row.purchase_profile,
                "business_name": row.business_name,
                "gender": row.gender,
                "birth_date": row.birth_date.strftime("%d/%m/%Y") if row.birth_date else None
            })

    # Sort by creation date
    all_registrations.sort(key=lambda x: x["created_at"] or "", reverse=True)

    if format.lower() == "excel":
        return await export_to_excel(all_registrations)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported export format. Use 'excel'"
        )


async def export_to_excel(registrations):
    """Export registrations to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cadastros"

    # Define headers
    headers = [
        "Tipo", "Nome/Razão Social", "Documento", "E-mail", "Telefone",
        "Data Cadastro", "Tipo Negócio", "Nome Contato", "Função Contato",
        "Perfil Compra", "Nome Negócio (CPF)", "Gênero", "Data Nascimento",
        "CEP", "Endereço", "Bairro", "Cidade", "Estado"
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add data
    for row_num, reg in enumerate(registrations, 2):
        # ws.cell(row=row_num, column=1, value=str(reg["id"]))
        ws.cell(row=row_num, column=1, value=reg["type"])
        ws.cell(row=row_num, column=2, value=reg["name"])
        ws.cell(row=row_num, column=3, value=reg["document"])
        ws.cell(row=row_num, column=4, value=reg["email"])
        ws.cell(row=row_num, column=5, value=reg["phone"])
        ws.cell(row=row_num, column=6, value=reg["created_at"])
        ws.cell(row=row_num, column=7, value=reg["business_type"])
        ws.cell(row=row_num, column=8, value=reg["contact_name"])
        ws.cell(row=row_num, column=9, value=reg["contact_role"])
        ws.cell(row=row_num, column=10, value=reg["purchase_profile"])
        ws.cell(row=row_num, column=11, value=reg["business_name"])
        ws.cell(row=row_num, column=12, value=reg["gender"])
        ws.cell(row=row_num, column=13, value=reg["birth_date"])
        ws.cell(row=row_num, column=14, value=reg["cep"])
        ws.cell(row=row_num, column=15, value=reg["endereco"])
        ws.cell(row=row_num, column=16, value=reg["bairro"])
        ws.cell(row=row_num, column=17, value=reg["cidade"])
        ws.cell(row=row_num, column=18, value=reg["estado"])

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        max_length = 0
        for row_num in range(1, len(registrations) + 2):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Create in-memory file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return as streaming response
    return StreamingResponse(
        io.BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cadastros.xlsx"}
    )


@router.get("/dashboard/stats")
async def get_dashboard_stats(
    request: Request,
    db: AsyncSession = Depends(get_database)
):
    """Get dashboard statistics for admin."""
    # Try to get token from Authorization header first
    token = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
    else:
        # Fall back to cookie
        token = request.cookies.get("admin_token")

    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required"
        )

    # Verify token
    payload = verify_token(token)
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication token"
        )

    # Check if user is admin
    if payload.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    # Get statistics
    cnpj_count = await db.execute(
        select(func.count()).select_from(CNPJRegistration)
    )
    total_cnpj = cnpj_count.scalar()

    cpf_count = await db.execute(
        select(func.count()).select_from(CPFRegistration)
    )
    total_cpf = cpf_count.scalar()

    org_count = await db.execute(
        select(func.count()).select_from(Organization)
    )
    total_orgs = org_count.scalar()

    user_count = await db.execute(
        select(func.count()).select_from(User)
    )
    total_users = user_count.scalar()

    # Get recent registrations (last 10)
    recent_cnpj = await db.execute(
        select(CNPJRegistration.id, CNPJRegistration.razao_social.label("name"),
               CNPJRegistration.cnpj.label("document"), CNPJRegistration.email,
               CNPJRegistration.celular.label("phone"), CNPJRegistration.created_at)
        .order_by(CNPJRegistration.created_at.desc())
        .limit(5)
    )
    recent_cpf = await db.execute(
        select(CPFRegistration.id, CPFRegistration.nome_completo.label("name"),
               CPFRegistration.cpf.label("document"), CPFRegistration.email,
               CPFRegistration.celular.label("phone"), CPFRegistration.created_at)
        .order_by(CPFRegistration.created_at.desc())
        .limit(5)
    )

    recent_registrations = []

    # Combine and sort recent registrations
    for reg in recent_cnpj:
        recent_registrations.append({
            "id": reg.id,
            "name": reg.name,
            "document": reg.document,
            "email": reg.email,
            "phone": reg.phone,
            "type": "CNPJ",
            "created_at": reg.created_at.isoformat()
        })

    for reg in recent_cpf:
        recent_registrations.append({
            "id": reg.id,
            "name": reg.name,
            "document": reg.document,
            "email": reg.email,
            "phone": reg.phone,
            "type": "CPF",
            "created_at": reg.created_at.isoformat()
        })

    # Sort by creation date (most recent first)
    recent_registrations.sort(key=lambda x: x["created_at"], reverse=True)
    recent_registrations = recent_registrations[:10]

    return {
        "total_cnpj_registrations": total_cnpj,
        "total_cpf_registrations": total_cpf,
        "total_organizations": total_orgs,
        "total_users": total_users,
        "recent_registrations": recent_registrations
    }


@router.get("/registrations")
async def get_registrations(
    request: Request,
    page: int = 1,
    limit: int = 20,
    registration_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_database),
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get paginated list of registrations for admin."""
    # Verify token and admin role
    token = None
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
    else:
        # Fall back to cookie
        token = request.cookies.get("admin_token")

    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required"
        )
    
    payload = verify_token(credentials.credentials)
    if not payload or payload.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    offset = (page - 1) * limit

    # Build queries for CNPJ and CPF registrations
    cnpj_query = select(
        CNPJRegistration.id,
        CNPJRegistration.razao_social.label("name"),
        CNPJRegistration.email,
        CNPJRegistration.celular.label("phone"),
        CNPJRegistration.created_at,
        literal("CNPJ").label("type")
    )

    cpf_query = select(
        CPFRegistration.id,
        CPFRegistration.nome_completo.label("name"),
        CPFRegistration.email,
        CPFRegistration.celular.label("phone"),
        CPFRegistration.created_at,
        literal("CPF").label("type")
    )

    # Apply filters
    if registration_type:
        if registration_type == "CNPJ":
            cpf_query = None
        elif registration_type == "CPF":
            cnpj_query = None

    if date_from:
        date_from_dt = parse_date_filter(date_from)
        if cnpj_query is not None:
            cnpj_query = cnpj_query.where(CNPJRegistration.created_at >= date_from_dt)
        if cpf_query is not None:
            cpf_query = cpf_query.where(CPFRegistration.created_at >= date_from_dt)

    if date_to:
        date_to_dt = parse_date_filter(date_to)
        # Set to end of day for date_to
        date_to_dt = datetime.combine(date_to_dt.date(), time.max)
        if cnpj_query is not None:
            cnpj_query = cnpj_query.where(CNPJRegistration.created_at <= date_to_dt)
        if cpf_query is not None:
            cpf_query = cpf_query.where(CPFRegistration.created_at <= date_to_dt)

    if search:
        search_term = f"%{search}%"
        if cnpj_query is not None:
            cnpj_query = cnpj_query.where(
                (CNPJRegistration.razao_social.ilike(search_term)) |
                (CNPJRegistration.email.ilike(search_term))
            )
        if cpf_query is not None:
            cpf_query = cpf_query.where(
                (CPFRegistration.nome_completo.ilike(search_term)) |
                (CPFRegistration.email.ilike(search_term))
            )

    # Execute queries and combine results
    all_registrations = []

    if cnpj_query is not None:
        cnpj_result = await db.execute(cnpj_query.offset(offset).limit(limit))
        for row in cnpj_result:
            all_registrations.append({
                "id": row.id,
                "name": row.name,
                "email": row.email,
                "phone": row.phone,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "type": row.type
            })

    if cpf_query is not None:
        cpf_result = await db.execute(cpf_query.offset(offset).limit(limit))
        for row in cpf_result:
            all_registrations.append({
                "id": row.id,
                "name": row.name,
                "email": row.email,
                "phone": row.phone,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "type": row.type
            })

    # Sort combined results by creation date
    all_registrations.sort(key=lambda x: x["created_at"], reverse=True)

    # Apply pagination to combined results
    total_registrations = len(all_registrations)
    paginated_registrations = all_registrations[offset:offset + limit]

    # Calculate pagination info
    total_pages = (total_registrations + limit - 1) // limit

    return {
        "registrations": paginated_registrations,
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total_registrations,
            "total_pages": total_pages,
            "has_prev": page > 1,
            "has_next": page < total_pages,
            "prev_num": page - 1 if page > 1 else None,
            "next_num": page + 1 if page < total_pages else None
        }
    }


@router.get("/registrations/{registration_id}")
async def get_registration_detail(
    registration_id: str,
    registration_type: str,
    db: AsyncSession = Depends(get_database),
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get detailed information about a specific registration."""
    # Verify token and admin role
    payload = verify_token(credentials.credentials)
    if not payload or payload.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    if registration_type.upper() == "CNPJ":
        result = await db.execute(
            select(CNPJRegistration).where(CNPJRegistration.id == registration_id)
        )
        registration = result.scalar_one_or_none()
    elif registration_type.upper() == "CPF":
        result = await db.execute(
            select(CPFRegistration).where(CPFRegistration.id == registration_id)
        )
        registration = result.scalar_one_or_none()
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid registration type"
        )

    if not registration:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Registration not found"
        )

    return registration


@router.delete("/registrations/{registration_id}")
async def delete_registration(
    registration_id: str,
    registration_type: str,
    db: AsyncSession = Depends(get_database),
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a registration (admin only)."""
    # Verify token and admin role
    payload = verify_token(credentials.credentials)
    if not payload or payload.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    if registration_type.upper() == "CNPJ":
        result = await db.execute(
            select(CNPJRegistration).where(CNPJRegistration.id == registration_id)
        )
        registration = result.scalar_one_or_none()
        if registration:
            await db.delete(registration)
            await db.commit()
    elif registration_type.upper() == "CPF":
        result = await db.execute(
            select(CPFRegistration).where(CPFRegistration.id == registration_id)
        )
        registration = result.scalar_one_or_none()
        if registration:
            await db.delete(registration)
            await db.commit()
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid registration type"
        )

    if not registration:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Registration not found"
        )

    return {"message": "Registration deleted successfully"}